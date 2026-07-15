"""
Ponto de entrada (Main Window) da Arquitetura do PySide6.
Controla o orquestrador dos sinais, estado e a barra superior do programa.
"""
import os
import sys
import numpy as np
import scipy.sparse as sparse
from scipy.io import mmwrite

from PySide6.QtWidgets import (QApplication, QMainWindow, QHBoxLayout, QWidget,
                               QMenu, QMessageBox, QFileDialog)
from PySide6.QtGui import QKeySequence, QShortcut
from PySide6.QtCore import QThread, Qt, Slot

import visualization
from .result_dataclass import ResultadoMEF
from .sidebar import Sidebar
from .viewer import ViewerWidget
from .visualizer import VisualizerManager
from .worker import TrabalhadorCalculo
from .styles import get_stylesheet
from .utils import formatar, garantir_extensao, detalhes_formato


class JanelaMEF(QMainWindow):
    """
    Janela Principal do TerraMEF. Instancia o Painel e o Visualizador e liga
    as interações do utilizador aos dados do trabalhador MEF.
    """

    def __init__(self) -> None:
        """Inicializa a UI e todas as suas variáveis de estado globais."""
        super().__init__()
        self.setWindowTitle("TerraMEF | Simulador de aterramento")
        self.resize(1420, 860)

        # Variáveis de Estado (Memória)
        self.resultado = None
        self.malha = None
        self.modo = "superficie"
        self.eixo_corte = "y"

        # Controlo de Thread
        self.thread = None
        self.trabalhador = None
        self.atalhos = []

        # Arranque da janela modular
        self._criar_interface()
        self.setStyleSheet(get_stylesheet())
        self._criar_atalhos()
        self._criar_barra_exportacao()
        self._conectar_sinais()

    def _criar_interface(self) -> None:
        """Monta o Layout Geral (Esquerda e Direita)."""
        raiz = QWidget()
        self.setCentralWidget(raiz)
        layout = QHBoxLayout(raiz)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # 1. Instanciar Sidebar Esquerda
        self.sidebar = Sidebar(self)
        layout.addWidget(self.sidebar)

        # 2. Instanciar Viewer 3D à direita
        self.viewer = ViewerWidget(self)
        layout.addWidget(self.viewer, 1)

        # 3. Preparar Gestor da renderização gráfica
        self.visualizer_manager = VisualizerManager(self.viewer.visualizador)

        # Ocultar painéis do modo corte, pois começa no modo de "superficie 3D"
        self.sidebar.label_orientacao_corte.setVisible(False)
        for botao in self.sidebar.botoes_eixos:
            botao.setVisible(False)
        self.sidebar.label_posicao_plano.setVisible(False)
        self.sidebar.slider_corte.setVisible(False)
        self.sidebar.posicao_corte.setVisible(False)
        self.sidebar.intervalo_corte.setVisible(False)
        self.sidebar.botao_repor_corte.setVisible(False)
        self.sidebar.mostrar_contornos.setVisible(False)
        self.sidebar.mostrar_setas.setVisible(False)
        self.viewer.navegacao_corte.setVisible(False)

        # Desativar (Dim)
        self.sidebar.slider_corte.setEnabled(False)
        self.sidebar.posicao_corte.setEnabled(False)
        self.sidebar.mostrar_contornos.setEnabled(False)
        self.sidebar.mostrar_setas.setEnabled(False)

    def _conectar_sinais(self) -> None:
        """Liga as ações dos botões (clicks) às funções internas."""
        self.sidebar.botao_calcular.clicked.connect(self.calcular)
        self.sidebar.botao_superficie.clicked.connect(lambda: self.definir_modo("superficie"))
        self.sidebar.botao_corte.clicked.connect(lambda: self.definir_modo("corte"))

        for botao in self.sidebar.botoes_eixos:
            botao.clicked.connect(lambda marcado=False, valor=botao.property("eixo"): self.definir_eixo(valor))

        self.sidebar.slider_corte.valueChanged.connect(self._slider_corte_alterado)
        self.sidebar.posicao_corte.valueChanged.connect(self._posicao_corte_alterada)
        self.sidebar.botao_repor_corte.clicked.connect(self.repor_corte)
        self.sidebar.mostrar_contornos.toggled.connect(self.atualizar_visualizacao)
        self.sidebar.mostrar_setas.toggled.connect(self.atualizar_visualizacao)
        self.sidebar.mostrar_malha.toggled.connect(self.atualizar_visualizacao)

        self.viewer.botao_repor.clicked.connect(self.repor_camara)
        for botao in self.viewer.botoes_navegacao:
            botao.clicked.connect(lambda marcado=False, valor=botao.property("passo"): self.mover_corte(valor))

    def _criar_atalhos(self) -> None:
        """Inicializa os atalhos de teclado (ex: 'A' e 'D' para mover o plano)."""
        for tecla, acao in (
                ("A", lambda: self.mover_corte(-1)),
                ("D", lambda: self.mover_corte(1)),
                ("Shift+A", lambda: self.mover_corte(-10)),
                ("Shift+D", lambda: self.mover_corte(10)),
                ("Left", lambda: self.mover_corte(-1)),
                ("Right", lambda: self.mover_corte(1)),
                ("R", self.repor_camara),
                ("P", self.guardar_imagem),
        ):
            atalho = QShortcut(QKeySequence(tecla), self)
            atalho.setContext(Qt.ShortcutContext.ApplicationShortcut)
            atalho.activated.connect(acao)
            self.atalhos.append(atalho)

    def _criar_barra_exportacao(self) -> None:
        """Monta o menu principal superior para guardar os resultados em ficheiros."""
        barra = self.menuBar()
        barra.setNativeMenuBar(False)
        menu_exportar = QMenu("Exportar", self)

        menu_exportar.addAction("Imagem da vista atual...", self.guardar_imagem)
        menu_exportar.addSeparator()
        menu_exportar.addAction("Livro Excel completo (.xlsx)...", self.exportar_excel)
        menu_exportar.addSeparator()

        # Sub-menu de Matrizes
        menu_rigidez = menu_exportar.addMenu("Matriz de rigidez")
        menu_rigidez.addAction("NPZ esparso (.npz)...", self.exportar_matriz_rigidez)
        menu_rigidez.addAction("Matrix Market (.mtx)...", self.exportar_matriz_rigidez_mtx)
        menu_rigidez.addAction("Texto tabulado (.txt)...", self.exportar_matriz_rigidez_txt)

        # Sub-menu de Nodos
        menu_potenciais = menu_exportar.addMenu("Potenciais por no")
        menu_potenciais.addAction("CSV (.csv)...", self.exportar_potenciais)
        menu_potenciais.addAction("Texto tabulado (.txt)...", lambda: self.exportar_potenciais("txt"))
        menu_potenciais.addAction("TSV (.tsv)...", lambda: self.exportar_potenciais("tsv"))

        # Sub-menu de Campo Elétrico
        menu_campo = menu_exportar.addMenu("Campo eletrico")
        menu_campo.addAction("CSV (.csv)...", self.exportar_campo_eletrico)
        menu_campo.addAction("Texto tabulado (.txt)...", lambda: self.exportar_campo_eletrico("txt"))
        menu_campo.addAction("TSV (.tsv)...", lambda: self.exportar_campo_eletrico("tsv"))

        # Sub-menu de Conectividades
        menu_elementos = menu_exportar.addMenu("Elementos")
        menu_elementos.addAction("CSV (.csv)...", self.exportar_elementos)
        menu_elementos.addAction("Texto tabulado (.txt)...", lambda: self.exportar_elementos("txt"))
        menu_elementos.addAction("TSV (.tsv)...", lambda: self.exportar_elementos("tsv"))

        barra.addMenu(menu_exportar)

    def calcular(self) -> None:
        """
        Despoleta a QThread para calcular o MEF quando se clica no botão principal.
        Bloqueia o botão até estar finalizado.
        """
        pontos = self.sidebar.caminho_pontos.text()
        elementos = self.sidebar.caminho_elementos.text()

        if not os.path.isfile(pontos) or not os.path.isfile(elementos):
            QMessageBox.warning(self, "Ficheiros em falta", "Selecione ficheiros validos para arrancar.")
            return

        self.sidebar.botao_calcular.setEnabled(False)
        self.mostrar_mensagem("A preparar o calculo...")

        self.thread = QThread(self)
        self.trabalhador = TrabalhadorCalculo(pontos, elementos)
        self.trabalhador.moveToThread(self.thread)

        # Engatar os sinais PySide do Trabalhador e da Thread
        self.thread.started.connect(self.trabalhador.executar)
        self.trabalhador.estado.connect(self.mostrar_mensagem)
        self.trabalhador.concluido.connect(self._calculo_concluido)
        self.trabalhador.falhou.connect(self._calculo_falhou)
        self.trabalhador.concluido.connect(self.thread.quit)
        self.trabalhador.falhou.connect(self.thread.quit)
        self.thread.finished.connect(self._limpar_thread)

        self.thread.start()  # Lança o processamento

    @Slot(object)
    def _calculo_concluido(self, resultado: ResultadoMEF) -> None:
        """Atuador chamado em sucesso: Recebe e aplica os dados processados."""
        self.resultado = resultado
        # Constrói o objeto tridimensional (UnstructuredGrid)
        self.malha = visualization.preparar_malha_vtk(resultado.nos, resultado.elementos)
        self.malha.point_data["Potencial (V)"] = resultado.potenciais
        self.malha.cell_data["Campo eletrico (V/m)"] = resultado.gradientes

        self._atualizar_indicadores()
        self._configurar_limites_corte()

        self.mostrar_mensagem("Calculo concluido. Use A/D, as setas ou os controlos laterais para mover o corte.")
        self.atualizar_visualizacao(repor_vista=True)  # Acorda o render final

    @Slot(str)
    def _calculo_falhou(self, detalhe: str) -> None:
        """Atuador chamado em falha."""
        self.mostrar_mensagem("O calculo nao foi concluido.")
        QMessageBox.critical(self, "Erro no calculo", detalhe.splitlines()[-1])

    def _limpar_thread(self) -> None:
        """Desbloqueia os botões e recicla memória."""
        self.sidebar.botao_calcular.setEnabled(True)
        if self.trabalhador is not None:
            self.trabalhador.deleteLater()
        if self.thread is not None:
            self.thread.deleteLater()
        self.trabalhador = None
        self.thread = None

    def _atualizar_indicadores(self) -> None:
        """Preenche as caixas de resultado (Resistência, Ohms, etc.)."""
        assert self.resultado is not None
        delta_v = max(self.resultado.condicoes.values()) - min(self.resultado.condicoes.values())
        corrente = delta_v / self.resultado.resistencia if self.resultado.resistencia else None

        valores = {
            "resistencia": formatar(self.resultado.resistencia, "ohm"),
            "corrente": formatar(corrente, "A"),
            "nos": str(len(self.resultado.nos)),
            "elementos": str(len(self.resultado.elementos)),
        }
        nomes = {"resistencia": "Resistencia", "corrente": "Corrente",
                 "nos": "Nos", "elementos": "Elementos"}

        for chave, valor in valores.items():
            self.sidebar.indicadores[chave].setText(f"<span>{nomes[chave]}</span><br><b>{valor}</b>")

    def definir_modo(self, modo: str) -> None:
        """Alterna a renderização gráfica entre Corte por Raios X ou Caixa Fechada 3D."""
        self.modo = modo
        ativo = modo == "corte"

        # Acender painéis do modo de Corte e apagar no modo Superficie
        self.sidebar.label_orientacao_corte.setVisible(ativo)
        for botao in self.sidebar.botoes_eixos:
            botao.setVisible(ativo)
            botao.setEnabled(ativo)
        self.sidebar.label_posicao_plano.setVisible(ativo)
        self.sidebar.slider_corte.setVisible(ativo)
        self.sidebar.posicao_corte.setVisible(ativo)
        self.sidebar.intervalo_corte.setVisible(ativo)
        self.sidebar.botao_repor_corte.setVisible(ativo)
        self.sidebar.mostrar_contornos.setVisible(ativo)
        self.sidebar.mostrar_setas.setVisible(ativo)
        self.sidebar.slider_corte.setEnabled(ativo)
        self.sidebar.posicao_corte.setEnabled(ativo)
        self.sidebar.mostrar_contornos.setEnabled(ativo)
        self.sidebar.mostrar_setas.setEnabled(ativo)
        self.viewer.navegacao_corte.setVisible(ativo)

        self.viewer.titulo_vista.setText("Corte interativo" if modo == "corte" else "Vista 3D do potencial")
        self.atualizar_visualizacao(repor_vista=True)

    def definir_eixo(self, eixo: str) -> None:
        """Roda o corte para plano X, Y, ou Z."""
        self.eixo_corte = eixo
        self._configurar_limites_corte()
        self.atualizar_visualizacao(repor_vista=True)

    def _configurar_limites_corte(self) -> None:
        """Ajusta o alcance máximo dos sliders de forma a não sair dos limites da malha física."""
        if self.malha is None: return
        limites = self.malha.bounds
        eixo_indice = {"x": 0, "y": 1, "z": 2}[self.eixo_corte]

        minimo, maximo = limites[2 * eixo_indice], limites[2 * eixo_indice + 1]
        margem = self._margem_corte(minimo, maximo)
        minimo_util, maximo_util = minimo + margem, maximo - margem

        valor = minimo_util + (
                    maximo_util - minimo_util) * self.sidebar.slider_corte.value() / self.sidebar.slider_corte.maximum()

        # Impede ecos infindos de Sinais Qt ao programar o valor
        self.sidebar.posicao_corte.blockSignals(True)
        self.sidebar.posicao_corte.setRange(minimo_util, maximo_util)
        self.sidebar.posicao_corte.setSingleStep((maximo_util - minimo_util) / 1000)
        self.sidebar.posicao_corte.setValue(valor)
        self.sidebar.posicao_corte.blockSignals(False)

        self.sidebar.intervalo_corte.setText(f"Limites: {minimo_util:.3f} a {maximo_util:.3f} m")
        self._atualizar_indicador_posicao(valor)

    @staticmethod
    def _margem_corte(minimo: float, maximo: float) -> float:
        """Mantém a fatia estritamente dentro da malha para não renderizar Vazio."""
        return max((maximo - minimo) * 1e-6, 1e-9)

    def _slider_corte_alterado(self) -> None:
        """Atuador ao agarrar na barra analógica de profundidade."""
        if self.malha is None: return
        posicao = self._origem_corte()[{"x": 0, "y": 1, "z": 2}[self.eixo_corte]]

        self.sidebar.posicao_corte.blockSignals(True)
        self.sidebar.posicao_corte.setValue(posicao)
        self.sidebar.posicao_corte.blockSignals(False)

        self._atualizar_indicador_posicao(posicao)
        self.atualizar_visualizacao()

    def _posicao_corte_alterada(self, posicao: float) -> None:
        """Atuador de introdução numérica do valor exato em Metros do corte."""
        if self.malha is None: return
        limites = self.malha.bounds
        indice = {"x": 0, "y": 1, "z": 2}[self.eixo_corte]
        minimo, maximo = limites[2 * indice], limites[2 * indice + 1]
        margem = self._margem_corte(minimo, maximo)
        minimo_util, maximo_util = minimo + margem, maximo - margem

        posicao = float(np.clip(posicao, minimo_util, maximo_util))
        proporcao = (posicao - minimo_util) / (maximo_util - minimo_util) if maximo_util > minimo_util else 0

        self.sidebar.slider_corte.blockSignals(True)
        self.sidebar.slider_corte.setValue(round(proporcao * self.sidebar.slider_corte.maximum()))
        self.sidebar.slider_corte.blockSignals(False)

        self._atualizar_indicador_posicao(posicao)
        self.atualizar_visualizacao()

    def _atualizar_indicador_posicao(self, posicao: float) -> None:
        """Edita o overlay de letras no Canvas 3D."""
        self.viewer.posicao_lateral.setText(f"{self.eixo_corte.upper()}\n{posicao:.2f} m")

    def mover_corte(self, passos: int) -> None:
        """Atuador de teclas. Movimenta o slider X passos."""
        if self.malha is None or self.modo != "corte": return
        novo_valor = max(self.sidebar.slider_corte.minimum(),
                         min(self.sidebar.slider_corte.maximum(), self.sidebar.slider_corte.value() + passos))
        self.sidebar.slider_corte.setValue(novo_valor)

    def repor_corte(self) -> None:
        """Limpa as opções para o meio, útil em malhas gigantes onde perdes de vista a fatia."""
        if self.malha is None: return
        self.sidebar.slider_corte.setValue(
            (self.sidebar.slider_corte.minimum() + self.sidebar.slider_corte.maximum()) // 2)
        self.mostrar_mensagem("Corte reposto ao centro, sem alterar a camara.")

    def _origem_corte(self) -> tuple[float, float, float]:
        """Calcula onde a lâmina de renderização irá passar espacialmente."""
        assert self.malha is not None
        limites = self.malha.bounds
        eixo_indice = {"x": 0, "y": 1, "z": 2}[self.eixo_corte]
        minimo, maximo = limites[2 * eixo_indice], limites[2 * eixo_indice + 1]
        margem = self._margem_corte(minimo, maximo)
        minimo_util, maximo_util = minimo + margem, maximo - margem

        valor = minimo_util + (
                    maximo_util - minimo_util) * self.sidebar.slider_corte.value() / self.sidebar.slider_corte.maximum()
        origem = list(self.malha.center)
        origem[eixo_indice] = valor
        return tuple(origem)

    @Slot()
    def atualizar_visualizacao(self, repor_vista: bool = False) -> None:
        """Chama a entidade Externa 'VisualizerManager' para renderizar os visuais 3D."""
        origem = self._origem_corte() if self.modo == "corte" and self.malha is not None else None

        self.visualizer_manager.atualizar_visualizacao(
            malha=self.malha,
            resultado=self.resultado,
            modo=self.modo,
            eixo_corte=self.eixo_corte,
            origem_corte=origem,
            mostrar_malha=self.sidebar.mostrar_malha.isChecked(),
            mostrar_contornos=self.sidebar.mostrar_contornos.isChecked(),
            mostrar_setas=self.sidebar.mostrar_setas.isChecked(),
            repor_vista=repor_vista,
            mensagem_callback=self.mostrar_mensagem
        )

    def repor_camara(self) -> None:
        """Pede ao PyVista Manager para focar e centralizar o modelo."""
        self.visualizer_manager.repor_camara(self.modo, self.eixo_corte, mensagem_callback=self.mostrar_mensagem)

    def guardar_imagem(self) -> None:
        """Guarda o print da Viewport em formato .png."""
        if self.malha is None: return
        destino, _ = QFileDialog.getSaveFileName(self, "Guardar imagem", "resultado_mef.png", "Imagem PNG (*.png)")
        if destino:
            self.viewer.visualizador.screenshot(destino, transparent_background=False)
            self.mostrar_mensagem(f"Imagem guardada: {os.path.basename(destino)}")

    def _confirmar_resultado_para_exportar(self) -> bool:
        """Bloqueia a barra de ficheiros se o utilizador não tiver gerado nenhum resultado matemático."""
        if self.resultado is None:
            QMessageBox.information(self, "Sem resultado", "Calcule primeiro a simulacao.")
            return False
        return True

    # ==============================================================
    # EXPORTAÇÃO PARA FICHEIROS TEXTO E EXCEL (NumPy SaveText)
    # ==============================================================

    def exportar_matriz_rigidez(self) -> None:
        """Exporta Matriz de Condutividade esparsa (CSR) com formato Numpy nativo."""
        if not self._confirmar_resultado_para_exportar(): return
        destino, _ = QFileDialog.getSaveFileName(self, "Exportar matriz", "matriz_rigidez.npz", "Matriz (*.npz)")
        if destino:
            sparse.save_npz(garantir_extensao(destino, ".npz"), self.resultado.matriz_rigidez)
            self.mostrar_mensagem(f"Exportada: {os.path.basename(destino)}")

    def exportar_matriz_rigidez_mtx(self) -> None:
        """Exporta Matriz de Condutividade esparsa no standard 'Matrix Market'."""
        if not self._confirmar_resultado_para_exportar(): return
        destino, _ = QFileDialog.getSaveFileName(self, "Exportar matriz", "matriz_rigidez.mtx", "Matrix Market (*.mtx)")
        if destino:
            mmwrite(garantir_extensao(destino, ".mtx"), self.resultado.matriz_rigidez)
            self.mostrar_mensagem(f"Exportada: {os.path.basename(destino)}")

    def exportar_matriz_rigidez_txt(self) -> None:
        """Exporta Matriz de Rigidez Esparsa descompactada para Coordenadas Human-Readable (Linha | Coluna | Valor)."""
        if not self._confirmar_resultado_para_exportar(): return
        destino, _ = QFileDialog.getSaveFileName(self, "Exportar matriz", "matriz_rigidez.txt", "Texto (*.txt)")
        if destino:
            destino = garantir_extensao(destino, ".txt")
            matriz = self.resultado.matriz_rigidez.tocoo()
            np.savetxt(destino, np.column_stack((matriz.row, matriz.col, matriz.data)),
                       delimiter="\t", header="linha\tcoluna\tvalor", comments="", fmt=["%d", "%d", "%.12g"])
            self.mostrar_mensagem(f"Exportada: {os.path.basename(destino)}")

    def exportar_potenciais(self, formato: str = "csv") -> None:
        """Guarda mapa dos Nós, coordenadas espaciais, e as suas tensões [V]."""
        if not self._confirmar_resultado_para_exportar(): return
        ext, filtro, sep = detalhes_formato(formato)
        destino, _ = QFileDialog.getSaveFileName(self, "Exportar", f"potenciais{ext}", filtro)
        if destino:
            destino = garantir_extensao(destino, ext)
            np.savetxt(destino, np.column_stack(
                (np.arange(len(self.resultado.nos)), self.resultado.nos, self.resultado.potenciais)),
                       delimiter=sep, header=f"no{sep}x{sep}y{sep}z{sep}potencial_V", comments="")
            self.mostrar_mensagem(f"Exportado: {os.path.basename(destino)}")

    def exportar_campo_eletrico(self, formato: str = "csv") -> None:
        """Guarda mapa dos centroides de cada Tetraedro (Elemento Finito), e os vetores de campo [V/m]."""
        if not self._confirmar_resultado_para_exportar(): return
        ext, filtro, sep = detalhes_formato(formato)
        destino, _ = QFileDialog.getSaveFileName(self, "Exportar", f"campo_eletrico{ext}", filtro)
        if destino:
            destino = garantir_extensao(destino, ext)
            centros = self.resultado.nos[self.resultado.elementos].mean(axis=1)
            np.savetxt(destino, np.column_stack((np.arange(len(centros)), centros, self.resultado.gradientes)),
                       delimiter=sep, header=f"elemento{sep}x{sep}y{sep}z{sep}Ex{sep}Ey{sep}Ez", comments="")
            self.mostrar_mensagem(f"Exportado: {os.path.basename(destino)}")

    def exportar_elementos(self, formato: str = "csv") -> None:
        """Guarda Tabela de Conectividade TET4."""
        if not self._confirmar_resultado_para_exportar(): return
        ext, filtro, sep = detalhes_formato(formato)
        destino, _ = QFileDialog.getSaveFileName(self, "Exportar", f"elementos{ext}", filtro)
        if destino:
            destino = garantir_extensao(destino, ext)
            np.savetxt(destino, np.column_stack((np.arange(len(self.resultado.elementos)), self.resultado.elementos)),
                       fmt="%d", delimiter=sep, header=f"elemento{sep}no_0{sep}no_1{sep}no_2{sep}no_3", comments="")
            self.mostrar_mensagem(f"Exportado: {os.path.basename(destino)}")

    def exportar_excel(self) -> None:
        """
        Monta um arquivo .xlsx com várias abas separadas contendo Relatórios Globais, Nódos, Matrizes, etc.
        (Requer a biblioteca OpenPyXL).
        """
        if not self._confirmar_resultado_para_exportar(): return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            QMessageBox.warning(self, "Dependencia", "Instale o modulo openpyxl (pip install openpyxl).")
            return

        destino, _ = QFileDialog.getSaveFileName(self, "Exportar", "resultado.xlsx", "Excel (*.xlsx)")
        if not destino: return

        destino = garantir_extensao(destino, ".xlsx")
        livro = Workbook()
        livro.remove(livro.active)  # Remove aba 'Sheet' de origem

        cabecalho = PatternFill("solid", fgColor="0F766E")
        fonte_cabecalho = Font(color="FFFFFF", bold=True)

        def escrever_folha(nome, titulos, linhas):
            folha = livro.create_sheet(nome)
            folha.append(titulos)
            for c in folha[1]:
                c.fill, c.font = cabecalho, fonte_cabecalho
            folha.freeze_panes = "A2"
            for linha in linhas:
                folha.append(linha)
            for col, titulo in enumerate(titulos, start=1):
                folha.column_dimensions[chr(64 + col)].width = max(14, len(titulo) + 3)

        diferenca = max(self.resultado.condicoes.values()) - min(self.resultado.condicoes.values())
        corrente = diferenca / self.resultado.resistencia if self.resultado.resistencia else None

        # ABA 1: Relatório de Metadados
        escrever_folha("Resumo", ["medida", "valor"], [
            ("Resistencia (ohm)", self.resultado.resistencia),
            ("Corrente (A)", corrente),
            ("Numero de nos", len(self.resultado.nos)),
            ("Numero de elementos", len(self.resultado.elementos)),
            ("Potencial minimo (V)", float(np.min(self.resultado.potenciais))),
            ("Potencial maximo (V)", float(np.max(self.resultado.potenciais)))
        ])

        # ABA 2: Tensões (Volts)
        escrever_folha("Potenciais", ["no", "x", "y", "z", "potencial_V"],
                       ((int(i), *map(float, no), float(pot)) for i, (no, pot) in
                        enumerate(zip(self.resultado.nos, self.resultado.potenciais))))

        # ABA 3: Campo elétrico nos tetraedros
        centros = self.resultado.nos[self.resultado.elementos].mean(axis=1)
        escrever_folha("Campo", ["elemento", "x", "y", "z", "Ex", "Ey", "Ez"],
                       ((int(i), *map(float, centro), *map(float, campo)) for i, (centro, campo) in
                        enumerate(zip(centros, self.resultado.gradientes))))

        # ABA 4: Conectividade MEF
        escrever_folha("Elementos", ["elemento", "no_0", "no_1", "no_2", "no_3"],
                       ((int(i), *map(int, elemento)) for i, elemento in enumerate(self.resultado.elementos)))

        livro.save(destino)
        self.mostrar_mensagem(f"Exportado: {os.path.basename(destino)}")

    @Slot(str)
    def mostrar_mensagem(self, texto: str) -> None:
        """Altera os quadros brancos com os logs de status."""
        self.sidebar.estado.setText(texto)
        self.viewer.mensagem_topo.setText(texto)


def iniciar_interface() -> int:
    """Função lançadora do QApplication em Boot."""
    aplicacao = QApplication.instance() or QApplication(sys.argv)
    aplicacao.setApplicationName("TerraMEF")
    janela = JanelaMEF()
    janela.show()
    return aplicacao.exec()